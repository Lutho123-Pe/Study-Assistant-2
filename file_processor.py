import os
from PyPDF2 import PdfReader
import docx
import openpyxl
import pytesseract
from PIL import Image

class FileProcessor:
    def __init__(self, documents_dir='data/documents'):
        self.documents_dir = documents_dir

    def load_documents(self):
        documents = []
        for filename in os.listdir(self.documents_dir):
            filepath = os.path.join(self.documents_dir, filename)
            if filename.endswith('.txt'):
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                    documents.append(content)
            elif filename.endswith('.pdf'):
                content = self._read_pdf(filepath)
                if not content.strip():
                    print(f"Warning: PDF {filename} may have extraction issues.")
                documents.append(content)
            elif filename.endswith('.docx'):
                content = self._read_docx(filepath)
                if not content.strip():
                    print(f"Warning: DOCX {filename} may have extraction issues.")
                documents.append(content)
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                content = self._read_excel(filepath)
                if not content.strip():
                    print(f"Warning: Excel {filename} may have extraction issues.")
                documents.append(content)
            elif filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                content = self._read_image_text(filepath)
                if not content.strip():
                    print(f"Warning: Image {filename} may have extraction issues.")
                documents.append(content)
        return documents

    def _read_pdf(self, filepath):
        text = ''
        try:
            reader = PdfReader(filepath)
            for page in reader.pages:
                text += page.extract_text() + '\n'
        except Exception as e:
            text = f"Error reading PDF {filepath}: {e}"
        return text

    def _read_docx(self, filepath):
        text = ''
        try:
            doc = docx.Document(filepath)
            for para in doc.paragraphs:
                text += para.text + '\n'
        except Exception as e:
            text = f"Error reading DOCX {filepath}: {e}"
        return text

    def _read_excel(self, filepath):
        text = ''
        try:
            wb = openpyxl.load_workbook(filepath)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' '.join([str(cell) if cell is not None else '' for cell in row])
                    text += row_text + '\n'
        except Exception as e:
            text = f"Error reading Excel {filepath}: {e}"
        return text

    def _read_image_text(self, filepath):
        text = ''
        try:
            image = Image.open(filepath)
            text = pytesseract.image_to_string(image)
        except Exception as e:
            text = f"Error reading image {filepath}: {e}"
        return text

    def process_file(self, filepath):
        """Process a single file and return its text content"""
        if filepath.endswith('.txt'):
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
        elif filepath.endswith('.pdf'):
            return self._read_pdf(filepath)
        elif filepath.endswith('.docx'):
            return self._read_docx(filepath)
        elif filepath.endswith('.xlsx') or filepath.endswith('.xls'):
            return self._read_excel(filepath)
        elif filepath.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
            return self._read_image_text(filepath)
        else:
            return f"Unsupported file type: {filepath}"
